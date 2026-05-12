#!/usr/bin/env python3
"""
Build a window/door schedule XLSX from Archicad data via Tapir.

Called by `ac schedule`. Reads configuration from environment:
  AC_PORT                 — Archicad's HTTP port (required)
  AC_SCHEDULE_OUT         — output xlsx path (default: ~/Desktop/Window-Door-Schedule.xlsx)
  AC_SCHEDULE_MIN_FLOOR   — minimum story index, inclusive (default: 0 = ground)
  AC_SCHEDULE_MAX_FLOOR   — maximum story index, inclusive (default: 999)
  AC_SCHEDULE_TITLE_PROJECT, AC_SCHEDULE_TITLE_CLIENT, AC_SCHEDULE_TITLE_DRAWING_NO,
  AC_SCHEDULE_TITLE_REV   — optional title-block strings

Requires: openpyxl, Pillow (provided via `uv run --with`).
"""
import json, os, sys, math, urllib.request
from io import BytesIO

from PIL import Image, ImageDraw, ImageFont
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# ── Config ────────────────────────────────────────────────────────────────────

PORT      = int(os.environ.get("AC_PORT", "0"))
OUT_PATH  = os.environ.get("AC_SCHEDULE_OUT", os.path.expanduser("~/Desktop/Window-Door-Schedule.xlsx"))
MIN_FLOOR = int(os.environ.get("AC_SCHEDULE_MIN_FLOOR", "0"))
MAX_FLOOR = int(os.environ.get("AC_SCHEDULE_MAX_FLOOR", "999"))

PROJECT   = os.environ.get("AC_SCHEDULE_TITLE_PROJECT", "")
CLIENT    = os.environ.get("AC_SCHEDULE_TITLE_CLIENT", "")
DRAW_NO   = os.environ.get("AC_SCHEDULE_TITLE_DRAWING_NO", "")
REV       = os.environ.get("AC_SCHEDULE_TITLE_REV", "P01")


# ── Tapir client (simple) ─────────────────────────────────────────────────────

def tapir(cmd, params=None):
    payload = {"command": "API.ExecuteAddOnCommand", "parameters": {
        "addOnCommandId": {"commandNamespace": "TapirCommand", "commandName": cmd},
        "addOnCommandParameters": params or {},
    }}
    req = urllib.request.Request(
        f"http://127.0.0.1:{PORT}",
        data=json.dumps(payload).encode(),
        headers={"Content-Type": "application/json"},
    )
    res = json.loads(urllib.request.urlopen(req, timeout=120).read())
    if res.get("succeeded") is False:
        raise RuntimeError(f"Archicad error on {cmd}: {res.get('error')}")
    return res["result"]["addOnCommandResponse"]


def get_project_name():
    try:
        proj = tapir("GetProjectInfoFields")
        return next(
            (f.get("projectInfoValue") for f in proj.get("fields", [])
             if f.get("projectInfoId") == "ProjectName"), ""
        ) or ""
    except Exception:
        return ""


# ── Data fetching ─────────────────────────────────────────────────────────────

def fetch_filtered(etype):
    """Returns (elems, details) for elements of etype on floors [MIN_FLOOR..MAX_FLOOR]."""
    elems = tapir("GetElementsByType", {"elementType": etype}).get("elements", [])
    if not elems:
        return [], []
    det = tapir("GetDetailsOfElements", {"elements": elems}).get("detailsOfElements", [])
    keep, keep_det = [], []
    for e, d in zip(elems, det):
        fi = d.get("floorIndex", -99)
        if MIN_FLOOR <= fi <= MAX_FLOOR:
            keep.append(e); keep_det.append(d)
    return keep, keep_det


def gdl_for(elements):
    if not elements:
        return []
    res = tapir("GetGDLParametersOfElements", {"elements": elements})
    return [{p["name"]: p.get("value") for p in g.get("parameters", [])}
            for g in res.get("gdlParametersOfElements", [])]


# ── Type grouping ─────────────────────────────────────────────────────────────

def signature(detail, params):
    """Return a hashable type signature for grouping."""
    lib = detail.get("details", {}).get("libPart", {}).get("name", "?")
    w = round(params.get("ac_wallhole_width", 0) or 0, 3)
    h = round(params.get("ac_wallhole_height", 0) or 0, 3)
    return (lib, w, h)


def group_by_type(elements, details, gdls):
    """Returns a list of TypeGroup dicts:
       {sig, lib, w_mm, h_mm, ids: [str], idxs: [int], rep_params: dict}
    Sorted by descending count, with stable secondary by lib name then size.
    """
    groups = {}
    for i, (e, d, p) in enumerate(zip(elements, details, gdls)):
        sig = signature(d, p)
        g = groups.setdefault(sig, {
            "sig": sig,
            "lib": sig[0],
            "w_mm": int(round(sig[1] * 1000)),
            "h_mm": int(round(sig[2] * 1000)),
            "ids": [],
            "idxs": [],
            "rep_params": p,    # first member's params is the representative
        })
        g["ids"].append(d.get("id", ""))
        g["idxs"].append(i)
    out = list(groups.values())
    out.sort(key=lambda g: (-len(g["ids"]), g["lib"], g["w_mm"], g["h_mm"]))
    return out


# ── Elevation drawing ─────────────────────────────────────────────────────────

# Render at 2-3x final display size for crisp lines when scaled in XLSX.
RENDER_PX_PER_M  = 280
MARGIN_TOP       = 36
MARGIN_BOTTOM    = 88
MARGIN_LEFT      = 110
MARGIN_RIGHT     = 36
DIM_TICK         = 7

# Lineweights tuned for architectural schedule conventions at 2x render resolution (display in xlsx ~50%).
LW_FRAME_OUT     = 3   # outer wallhole boundary
LW_FRAME_IN      = 2   # inner frame edge (glass-side)
LW_MULLION_OUT   = 2   # mullion / transom outer
LW_MULLION_IN    = 2   # mullion / transom inner
LW_GLASS         = 1   # glass-area inner outline
LW_DIM           = 1
LW_DASHED        = 1
LW_HANDLE        = 2

FONT_PATHS = [
    "/System/Library/Fonts/Helvetica.ttc",
    "/System/Library/Fonts/HelveticaNeue.ttc",
    "/Library/Fonts/Arial.ttf",
    "/System/Library/Fonts/Supplemental/Arial.ttf",
]


def _font(size):
    for p in FONT_PATHS:
        if os.path.exists(p):
            try:
                return ImageFont.truetype(p, size)
            except Exception:
                pass
    return ImageFont.load_default()


def _dashed_line(draw, p1, p2, color="black", width=1, dash=(9, 6)):
    """Draw a dashed straight line. dash = (on, off) px."""
    x1, y1 = p1; x2, y2 = p2
    dx, dy = x2 - x1, y2 - y1
    L = math.hypot(dx, dy)
    if L < 1: return
    ux, uy = dx / L, dy / L
    pos = 0.0
    drawing = True
    while pos < L:
        seg = dash[0] if drawing else dash[1]
        end = min(pos + seg, L)
        if drawing:
            sx, sy = x1 + ux*pos, y1 + uy*pos
            ex, ey = x1 + ux*end, y1 + uy*end
            draw.line([(sx, sy), (ex, ey)], fill=color, width=width)
        pos = end
        drawing = not drawing


def _canvas_for(width_mm, height_mm):
    """Return (img, draw, x0, y0, x1, y1, scale_px_per_mm)."""
    if width_mm <= 0 or height_mm <= 0:
        img = Image.new("RGB", (300, 380), "white")
        d = ImageDraw.Draw(img)
        d.text((20, 180), "(no dimensions)", fill="black", font=_font(20))
        return img, d, 0, 0, 0, 0, 0

    Wpx = max(80, int(round(width_mm / 1000 * RENDER_PX_PER_M)))
    Hpx = max(80, int(round(height_mm / 1000 * RENDER_PX_PER_M)))
    MAX_DIM = 720
    if Wpx > MAX_DIM or Hpx > MAX_DIM:
        s = MAX_DIM / max(Wpx, Hpx)
        Wpx, Hpx = int(Wpx*s), int(Hpx*s)

    canvas_w = Wpx + MARGIN_LEFT + MARGIN_RIGHT
    canvas_h = Hpx + MARGIN_TOP + MARGIN_BOTTOM
    img = Image.new("RGB", (canvas_w, canvas_h), "white")
    draw = ImageDraw.Draw(img)

    x0, y0 = MARGIN_LEFT, MARGIN_TOP
    x1, y1 = MARGIN_LEFT + Wpx, MARGIN_TOP + Hpx
    scale = Wpx / width_mm
    return img, draw, x0, y0, x1, y1, scale


def _draw_dims(draw, x0, y0, x1, y1, width_mm, height_mm, label_w=None, label_h=None,
               internal_heights_mm=None):
    """Draw width-below + height-left dimension lines with end ticks and labels.
    `label_w` / `label_h` override the default integer label.
    `internal_heights_mm` adds small ticks + labels along the right edge for sash heights.
    """
    font_dim = _font(20)
    font_int = _font(15)

    # Width dim
    dim_y = y1 + 28
    draw.line([(x0, dim_y), (x1, dim_y)], fill="black", width=LW_DIM)
    for x in (x0, x1):
        draw.line([(x-DIM_TICK, dim_y+DIM_TICK), (x+DIM_TICK, dim_y-DIM_TICK)],
                  fill="black", width=LW_DIM+1)
    s = label_w if label_w is not None else f"{width_mm:,}"
    bbox = draw.textbbox((0,0), s, font=font_dim)
    tw = bbox[2]-bbox[0]
    draw.text(((x0+x1)/2 - tw/2, dim_y + 6), s, fill="black", font=font_dim)

    # Height dim (left side)
    dim_x = x0 - 50
    draw.line([(dim_x, y0), (dim_x, y1)], fill="black", width=LW_DIM)
    for y in (y0, y1):
        draw.line([(dim_x-DIM_TICK, y+DIM_TICK), (dim_x+DIM_TICK, y-DIM_TICK)],
                  fill="black", width=LW_DIM+1)
    s = label_h if label_h is not None else f"{height_mm:,}"
    bbox = draw.textbbox((0,0), s, font=font_dim)
    tw, th = bbox[2]-bbox[0], bbox[3]-bbox[1]
    img = Image.new("RGBA", (tw+8, th+8), (255,255,255,0))
    d2 = ImageDraw.Draw(img)
    d2.text((4, 2), s, fill="black", font=font_dim)
    img = img.rotate(90, expand=True)
    # Paste through the parent draw's image — we can't paste from a sub-draw, so
    # callers must do this via the surrounding image. Workaround: receive image too.
    return img, dim_x  # caller handles paste


def _label_rotated(parent_img, text, x_center, y_center, font, color="black"):
    bbox = ImageDraw.Draw(parent_img).textbbox((0,0), text, font=font)
    tw, th = bbox[2]-bbox[0], bbox[3]-bbox[1]
    img = Image.new("RGBA", (tw+8, th+8), (255,255,255,0))
    ImageDraw.Draw(img).text((4, 2), text, fill=color, font=font)
    img = img.rotate(90, expand=True)
    parent_img.paste(img, (int(x_center - img.size[0]/2), int(y_center - img.size[1]/2)), img)


def _read_layout(params, total_width_mm, total_height_mm):
    """Determine the actual sash layout of a window, driven by `ifc_optypestr`.

    `ifc_optypestr` is set by Archicad to the actual layout the architect chose
    in Window Settings (e.g. 'Single', 'Double Vertical', 'Triple Vertical', etc.).
    The `gs_optype_NN` slots are unreliable — they hold library *template* defaults
    that exist whether or not the layout actually has that many sashes.

    Returns dict:
      {
        "rows": int,                # number of horizontal bands (top→bottom)
        "cols": int,                # number of columns within each band
        "panes": [                  # row-major list of panes, total = rows*cols
            {"fixed": bool, "label": str, "row": int, "col": int}, ...
        ]
      }
    """
    layout_str = (params.get("ifc_optypestr") or "").strip()
    layout_int = params.get("ifc_optype")
    s = layout_str.lower()

    # Heuristic mapping from common ifc_optypestr values to (rows, cols).
    # If unknown, fall back to a single pane.
    if "double vertical" in s:           rows, cols = 1, 2
    elif "triple vertical" in s:         rows, cols = 1, 3
    elif "double horizontal" in s:       rows, cols = 2, 1
    elif "triple horizontal" in s:       rows, cols = 3, 1
    elif "with transom" in s:            rows, cols = 2, 1   # transom above main pane
    elif "with sill" in s:               rows, cols = 2, 1   # main pane above sill light
    else:                                rows, cols = 1, 1   # 'Single' or anything else

    # Determine fixed/opening per pane. For now, use the first N gs_optype_NN values
    # mapped row-major. If a pane's optype is "Fixed Sash"/"Fixed Glass" or m=2/3,
    # mark it fixed; else it's operable.
    panes = []
    n = rows * cols
    for i in range(n):
        slot = i + 1
        opt   = params.get(f"gs_optype_{slot:02d}")
        opt_m = params.get(f"gs_optype_m_{slot:02d}")
        is_fixed = bool(opt_m in (2, 3) or (isinstance(opt, str) and "Fixed" in opt))
        # For 'Single' layout we don't trust any of this — that's just 1 operable pane
        # by default unless the user explicitly set it to Fixed Glass.
        if rows == 1 and cols == 1:
            # Heuristic: if gs_optype_01 is "Fixed Glass" (m=3) -> truly fixed.
            # Otherwise treat as operable (default user intent).
            is_fixed = (opt_m == 3) or (isinstance(opt, str) and "Fixed Glass" in opt)
        panes.append({
            "fixed": is_fixed,
            "label": opt or "Open",
            "row": i // cols,
            "col": i %  cols,
        })

    return {"rows": rows, "cols": cols, "panes": panes, "layout": layout_str or "Single"}


# ── Window drawing (gold-standard inspired) ───────────────────────────────────

def draw_window(width_mm, height_mm, params):
    img, draw, x0, y0, x1, y1, scale = _canvas_for(width_mm, height_mm)
    if scale == 0:
        return img
    font_dim = _font(20)
    font_int = _font(14)
    font_fl  = _font(12)

    frame_thk_mm = float(params.get("gs_frame_thk", 0.06) or 0.06) * 1000
    sash_thk_mm  = float(params.get("gs_sash_thk", 0.05) or 0.05) * 1000
    frame_thk_px = max(6, min(int(frame_thk_mm * scale), int((y1-y0)*0.18)))
    mullion_px   = max(5, min(int(sash_thk_mm * scale), 14))
    glass_inset  = max(3, frame_thk_px // 3)

    # Outer wallhole boundary
    draw.rectangle([x0, y0, x1, y1], outline="black", width=LW_FRAME_OUT)
    # Inner frame edge
    fx0, fy0 = x0 + frame_thk_px, y0 + frame_thk_px
    fx1, fy1 = x1 - frame_thk_px, y1 - frame_thk_px
    draw.rectangle([fx0, fy0, fx1, fy1], outline="black", width=LW_FRAME_IN)

    # ── Layout from ifc_optypestr (the real source of truth) ──
    L = _read_layout(params, width_mm, height_mm)
    rows, cols, panes = L["rows"], L["cols"], L["panes"]

    # Compute per-pane rectangles (row-major). Even split for now (Archicad's
    # actual sash heights/widths require deeper GDL inspection).
    avail_w = fx1 - fx0
    avail_h = fy1 - fy0
    pane_rects = []
    for r in range(rows):
        for c in range(cols):
            pane_x0 = fx0 + int(round(c       * avail_w / cols))
            pane_x1 = fx0 + int(round((c+1)   * avail_w / cols))
            pane_y0 = fy0 + int(round(r       * avail_h / rows))
            pane_y1 = fy0 + int(round((r+1)   * avail_h / rows))
            pane_rects.append((pane_x0, pane_y0, pane_x1, pane_y1, panes[r*cols + c]))

    # Draw mullions/transoms as thick parallel-line bands between panes
    if cols > 1:
        for c in range(1, cols):
            mx = fx0 + int(round(c * avail_w / cols))
            draw.line([(mx, fy0), (mx, fy1)], fill="black", width=LW_MULLION_OUT)
            draw.line([(mx + mullion_px, fy0), (mx + mullion_px, fy1)],
                      fill="black", width=LW_MULLION_IN)
    if rows > 1:
        for r in range(1, rows):
            my = fy0 + int(round(r * avail_h / rows))
            draw.line([(fx0, my), (fx1, my)], fill="black", width=LW_MULLION_OUT)
            draw.line([(fx0, my + mullion_px), (fx1, my + mullion_px)],
                      fill="black", width=LW_MULLION_IN)

    # Per-pane: glass-area outline + opening triangle (or FL label if fixed)
    for (px0, py0, px1, py1, pane) in pane_rects:
        # Account for mullion band thickness when adjacent
        if pane["col"] > 0:        px0 += mullion_px
        if pane["row"] > 0:        py0 += mullion_px
        if px1 - px0 < 12 or py1 - py0 < 12: continue

        gi = glass_inset
        gx0, gy0 = px0 + gi, py0 + gi
        gx1, gy1 = px1 - gi, py1 - gi
        if gx1 - gx0 < 8 or gy1 - gy0 < 8: continue
        draw.rectangle([gx0, gy0, gx1, gy1], outline="black", width=LW_GLASS)

        if pane["fixed"]:
            label = "FL"
            box_w, box_h = 28, 18
            bx1_, by0_ = gx1 - 4, gy0 + 4
            bx0_ = bx1_ - box_w
            by1_ = by0_ + box_h
            draw.rectangle([bx0_, by0_, bx1_, by1_], outline="black", width=1)
            tb = draw.textbbox((0,0), label, font=font_fl)
            tw_, th_ = tb[2]-tb[0], tb[3]-tb[1]
            draw.text((bx0_ + (box_w-tw_)/2, by0_ + (box_h-th_)/2 - 1), label,
                      fill="black", font=font_fl)
        else:
            # Opening triangle: apex at bottom-center (hopper / bottom-hung inward)
            cx = (gx0 + gx1) / 2
            apex_y = gy1 - 3
            top_l = (gx0 + 3, gy0 + 3)
            top_r = (gx1 - 3, gy0 + 3)
            _dashed_line(draw, (cx, apex_y), top_l, color="black", width=LW_DASHED, dash=(9,6))
            _dashed_line(draw, (cx, apex_y), top_r, color="black", width=LW_DASHED, dash=(9,6))

    # ── Dimensions ──
    dim_y = y1 + 30
    draw.line([(x0, dim_y), (x1, dim_y)], fill="black", width=LW_DIM)
    for x in (x0, x1):
        draw.line([(x-DIM_TICK, dim_y+DIM_TICK), (x+DIM_TICK, dim_y-DIM_TICK)],
                  fill="black", width=LW_DIM+1)
    s_w = f"{width_mm:,}"
    bbox = draw.textbbox((0,0), s_w, font=font_dim)
    tw = bbox[2]-bbox[0]
    draw.text(((x0+x1)/2 - tw/2, dim_y + 6), s_w, fill="black", font=font_dim)

    dim_x = x0 - 56
    draw.line([(dim_x, y0), (dim_x, y1)], fill="black", width=LW_DIM)
    for y in (y0, y1):
        draw.line([(dim_x-DIM_TICK, y+DIM_TICK), (dim_x+DIM_TICK, y-DIM_TICK)],
                  fill="black", width=LW_DIM+1)
    _label_rotated(img, f"{height_mm:,}", dim_x - 16, (y0+y1)/2, font_dim)

    return img


# ── Door drawing (single + double leaf, hinge-aware) ─────────────────────────

def draw_door(width_mm, height_mm, params):
    img, draw, x0, y0, x1, y1, scale = _canvas_for(width_mm, height_mm)
    if scale == 0:
        return img
    font_dim   = _font(20)
    font_int   = _font(15)

    leaf_w_mm  = float(params.get("ac_leaf_width", 0) or 0) * 1000
    leaf_h_mm  = float(params.get("ac_leaf_height", 0) or 0) * 1000
    frame_thk_mm = float(params.get("gs_frame_thk", 0.06) or 0.06) * 1000
    second_w_mm  = float(params.get("gs_SecondLeaf_w", 0) or 0) * 1000
    leaf_widths_str = params.get("gs_list_doorleafwidths", "") or ""
    # Format examples seen: '950', '1,172', '713 / 713', '910 / 510 / 510'
    leaf_widths = []
    for tok in leaf_widths_str.replace(",", "").split("/"):
        tok = tok.strip()
        if not tok: continue
        try:
            leaf_widths.append(int(round(float(tok))))
        except ValueError:
            pass
    hinge_side = (params.get("ac_OpeningSide") or "L").upper()  # 'L' or 'R'
    optype     = params.get("gs_leaf_optype", "Side Hung") or "Side Hung"
    is_double  = ("Double" in optype) or (len(leaf_widths) == 2) or ("Double" in (params.get("__lib_name") or ""))
    has_handle = bool(params.get("gs_handle_type")) and (params.get("gs_handle_type") != "None")

    # ── Outer wallhole ──
    draw.rectangle([x0, y0, x1, y1], outline="black", width=LW_FRAME_OUT)

    # ── Frame opening (inset by frame thickness on top & sides; bottom flush) ──
    frame_thk_px = max(5, min(int(frame_thk_mm * scale), int((y1-y0)*0.06)))
    fx0 = x0 + frame_thk_px
    fy0 = y0 + frame_thk_px
    fx1 = x1 - frame_thk_px
    fy1 = y1   # door reaches floor
    draw.rectangle([fx0, fy0, fx1, fy1], outline="black", width=LW_FRAME_IN)

    # ── Leaf rectangle(s) ──
    # Use ac_leaf_width / ac_leaf_height when available; else infer from gs_list_doorleafwidths.
    if leaf_h_mm <= 0:
        leaf_h_mm = height_mm - frame_thk_mm  # fallback
    leaf_h_px = int(round(leaf_h_mm * scale))
    leaf_top_y = fy1 - leaf_h_px

    if is_double:
        # Two equal leaves meeting at the centre
        each_w_mm = leaf_widths[0] if leaf_widths else (leaf_w_mm / 2 if leaf_w_mm else (width_mm - 2*frame_thk_mm) / 2)
        each_w_px = int(round(each_w_mm * scale))
        # Distribute leaves so the pair is centered between fx0 and fx1
        pair_w_px = each_w_px * 2
        left_x0 = fx0 + ((fx1 - fx0) - pair_w_px) // 2
        leaves = [
            (left_x0,            left_x0 + each_w_px,         "L"),  # left leaf, hinge-LEFT
            (left_x0 + each_w_px, left_x0 + 2*each_w_px,      "R"),  # right leaf, hinge-RIGHT
        ]
    else:
        each_w_mm = leaf_widths[0] if leaf_widths else (leaf_w_mm if leaf_w_mm else width_mm - 2*frame_thk_mm)
        each_w_px = int(round(each_w_mm * scale))
        # Centre horizontally inside the wallhole opening
        leaf_x0 = fx0 + ((fx1 - fx0) - each_w_px) // 2
        leaves = [(leaf_x0, leaf_x0 + each_w_px, hinge_side)]

    for (lx0, lx1, hinge) in leaves:
        # Leaf rectangle (slightly thicker to suggest the door slab)
        draw.rectangle([lx0, leaf_top_y, lx1, fy1], outline="black", width=LW_FRAME_IN)

        # Swing triangle: vertices at the two hinge corners (top & bottom) and at the
        # latch-edge mid-height. Apex is on hinge side.
        if hinge == "L":
            hx_top  = (lx0,        leaf_top_y)
            hx_bot  = (lx0,        fy1)
            apex    = (lx1,        (leaf_top_y + fy1) // 2)
        else:  # R
            hx_top  = (lx1,        leaf_top_y)
            hx_bot  = (lx1,        fy1)
            apex    = (lx0,        (leaf_top_y + fy1) // 2)
        # Two dashed diagonals: from each hinge corner to the apex
        _dashed_line(draw, hx_top, apex, color="black", width=LW_DASHED, dash=(9,6))
        _dashed_line(draw, hx_bot, apex, color="black", width=LW_DASHED, dash=(9,6))

        # Handle: small lever on the latch side at ~1000mm from floor
        if has_handle:
            handle_y = fy1 - int(round(1000 * scale))  # 1000mm from floor
            if hinge == "L":
                # Latch on right; handle pokes out to the LEFT (into the leaf)
                hx = lx1 - max(8, int(round(60 * scale)))
                lever_x = lx1 - max(28, int(round(140 * scale)))
            else:
                hx = lx0 + max(8, int(round(60 * scale)))
                lever_x = lx0 + max(28, int(round(140 * scale)))
            # Rosette (small circle)
            r = max(3, int(round(15 * scale)))
            draw.ellipse([hx-r, handle_y-r, hx+r, handle_y+r], outline="black", width=LW_HANDLE)
            # Lever (short horizontal line ending in a small dot)
            draw.line([(hx, handle_y), (lever_x, handle_y)], fill="black", width=LW_HANDLE)

    # ── Dimensions: S/O <total>, secondary leaf widths under it ──
    # Outer S/O
    dim_y = y1 + 32
    draw.line([(x0, dim_y), (x1, dim_y)], fill="black", width=LW_DIM)
    for x in (x0, x1):
        draw.line([(x-DIM_TICK, dim_y+DIM_TICK), (x+DIM_TICK, dim_y-DIM_TICK)],
                  fill="black", width=LW_DIM+1)
    sw = f"S/O {width_mm:,}"
    bbox = draw.textbbox((0,0), sw, font=font_dim); tw = bbox[2]-bbox[0]
    draw.text(((x0+x1)/2 - tw/2, dim_y + 6), sw, fill="black", font=font_dim)

    # Inner leaf-width dim line (only if leaf widths known)
    if leaves and leaf_widths:
        inner_dim_y = y1 + 60
        # Just label each leaf width below
        for (lx0, lx1, _), w_mm in zip(leaves, leaf_widths):
            s = f"{w_mm:,}"
            bbox = draw.textbbox((0,0), s, font=font_int); tw = bbox[2]-bbox[0]
            draw.text(((lx0+lx1)/2 - tw/2, inner_dim_y), s, fill="black", font=font_int)

    # Height S/O (left)
    dim_x = x0 - 56
    draw.line([(dim_x, y0), (dim_x, y1)], fill="black", width=LW_DIM)
    for y in (y0, y1):
        draw.line([(dim_x-DIM_TICK, y+DIM_TICK), (dim_x+DIM_TICK, y-DIM_TICK)],
                  fill="black", width=LW_DIM+1)
    _label_rotated(img, f"S/O {height_mm:,}", dim_x - 16, (y0+y1)/2, font_dim)

    return img


# ── Dispatcher (kept for compatibility) ───────────────────────────────────────

def draw_elevation(width_mm, height_mm, params, is_door=False):
    if is_door:
        return draw_door(width_mm, height_mm, params)
    return draw_window(width_mm, height_mm, params)


# ── Per-type data extraction ──────────────────────────────────────────────────

# Schedule table rows. Most pull from the gs_list_* GDL parameters that Archicad's
# Window/Door Settings dialog populates. Empty strings render as blank cells.
TABLE_ROWS = [
    ("Type",                      lambda g: g["__type_label"]),
    ("Library Object",            lambda g: g["lib"]),
    ("Structural Opening (W×H)", lambda g: f"{g['w_mm']:,}×{g['h_mm']:,}" if g['w_mm'] else "-"),
    ("Fire Rating",               lambda g: (g["rep_params"].get("gs_list_firerating") or "").strip() or "-"),
    ("Fire Exit",                 lambda g: ("☑" if g["rep_params"].get("gs_list_optype","").lower().find("exit")>=0 else "☐")),
    ("Operation",                 lambda g: (g["rep_params"].get("gs_list_optype") or "").strip() or "-"),
    ("Glazing",                   lambda g: (g["rep_params"].get("gs_list_glazing") or "").strip() or "-"),
    ("Material Specification",    lambda g: (g["rep_params"].get("gs_list_finish") or "").strip() or "-"),
    ("External Frame Finish",     lambda g: (g["rep_params"].get("gs_list_finish") or "").strip() or "-"),
    ("Internal Frame Finish",     lambda g: "-"),
    ("Manufacturer",              lambda g: (g["rep_params"].get("gs_list_manufacturer") or "").strip() or "-"),
    ("Acoustic Rating",           lambda g: (g["rep_params"].get("gs_list_acousticrating") or "").strip() or "-"),
    ("Notes",                     lambda g: (g["rep_params"].get("gs_list_note") or "").strip() or ""),
    ("Element IDs",               lambda g: _format_ids(g["ids"])),
    ("Quantity",                  lambda g: str(len(g["ids"]))),
]


def _id_key(s):
    """Sort key for Archicad IDs like 'W0.07' or 'W.1.02': split by digit-runs."""
    import re
    parts = re.split(r"(\d+)", s)
    return tuple(int(p) if p.isdigit() else p.lower() for p in parts)


def _format_ids(ids):
    """Dedup and sort Archicad element IDs. If many share the same id, show count."""
    from collections import Counter
    if not ids:
        return ""
    counts = Counter(ids)
    keys = sorted(counts.keys(), key=_id_key)
    parts = []
    for k in keys:
        n = counts[k]
        parts.append(f"{k} (×{n})" if n > 1 else k)
    return ", ".join(parts)


# ── XLSX construction ─────────────────────────────────────────────────────────

THIN  = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def build_sheet(wb, title, types):
    """Render one sheet: types-as-columns, gold-standard layout."""
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False

    # Section header (row 1)
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(2, 1 + len(types)))
    ws.cell(row=1, column=1).fill = HDR_FILL

    LABEL_COL = 1
    FIRST_TYPE_COL = 2

    # Label column setup
    ws.column_dimensions[get_column_letter(LABEL_COL)].width = 30

    # Image row (row 3) — tall
    IMG_ROW = 3
    ws.row_dimensions[IMG_ROW].height = 230  # tall enough for typical elevation
    ws.cell(row=IMG_ROW, column=LABEL_COL, value="Elevation View").font = Font(bold=True, size=11)
    ws.cell(row=IMG_ROW, column=LABEL_COL).alignment = Alignment(horizontal="left", vertical="center")

    # Each type column: image + data rows
    for ti, g in enumerate(types):
        col = FIRST_TYPE_COL + ti
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 22

        # Type label in the row 2 below header
        type_label = f"Type {ti+1:02d}"
        g["__type_label"] = type_label

        # ── Render elevation, embed ──
        elev = draw_elevation(g["w_mm"], g["h_mm"], g["rep_params"], is_door=("Door" in g["lib"]))
        bio = BytesIO()
        elev.save(bio, format="PNG", optimize=True)
        bio.seek(0)
        xl_img = XLImage(bio)
        # Scale to fit the column — Excel column widths are in "characters", roughly 7 px per unit
        target_w_px = 22 * 7  # ≈ 154 px column width
        scale = target_w_px / elev.size[0]
        xl_img.width = int(elev.size[0] * scale)
        xl_img.height = int(elev.size[1] * scale)
        # Anchor to the cell
        anchor_cell = f"{col_letter}{IMG_ROW}"
        ws.add_image(xl_img, anchor_cell)

        # If the scaled height exceeds row height, expand the row a bit
        needed_pts = xl_img.height * 0.75  # px → pt
        if needed_pts > ws.row_dimensions[IMG_ROW].height:
            ws.row_dimensions[IMG_ROW].height = min(needed_pts + 6, 320)

    # Data rows (start two rows below image row)
    DATA_START = IMG_ROW + 2
    for ri, (label, fn) in enumerate(TABLE_ROWS):
        r = DATA_START + ri
        # Bold-ish label rows
        is_bold = label in ("Type", "Quantity")
        cell = ws.cell(row=r, column=LABEL_COL, value=label)
        cell.font = Font(bold=is_bold, size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER
        if is_bold:
            cell.fill = HDR_FILL
        # Bigger row for IDs / Notes (potentially long content)
        if label in ("Element IDs", "Notes"):
            ws.row_dimensions[r].height = 60
        for ti, g in enumerate(types):
            col = FIRST_TYPE_COL + ti
            try:
                val = fn(g)
            except Exception:
                val = ""
            c = ws.cell(row=r, column=col, value=val)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.font = Font(size=10, bold=is_bold)
            c.border = BORDER
            if is_bold:
                c.fill = HDR_FILL

    # Freeze first column + top header rows
    ws.freeze_panes = ws.cell(row=DATA_START, column=FIRST_TYPE_COL)


def build_title_block(wb):
    """Tiny optional title sheet (in lieu of a layout-style title block we can't easily replicate)."""
    ws = wb.create_sheet("Title", 0)
    ws.sheet_view.showGridLines = False
    proj = PROJECT or get_project_name() or "(unsaved project)"
    rows = [
        ("Project", proj),
        ("Client", CLIENT or ""),
        ("Drawing", "External Door/Window Types"),
        ("Drawing No.", DRAW_NO or ""),
        ("Revision", REV),
        ("Generated by", "ac schedule (Tapir-fast)"),
    ]
    for i, (k, v) in enumerate(rows, start=2):
        ws.cell(row=i, column=2, value=k).font = Font(bold=True)
        ws.cell(row=i, column=3, value=v)
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 50
    ws.cell(row=1, column=2, value="External Door / Window Schedule").font = Font(bold=True, size=14)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if not PORT:
        print("AC_PORT not set", file=sys.stderr); sys.exit(1)

    print(f"Fetching windows on floors [{MIN_FLOOR}..{MAX_FLOOR}]...", file=sys.stderr)
    win, win_det = fetch_filtered("Window")
    print(f"  {len(win)} windows", file=sys.stderr)
    print(f"Fetching doors on floors [{MIN_FLOOR}..{MAX_FLOOR}]...", file=sys.stderr)
    dor, dor_det = fetch_filtered("Door")
    print(f"  {len(dor)} doors", file=sys.stderr)

    print("Fetching GDL parameters (batch)...", file=sys.stderr)
    win_gdl = gdl_for(win)
    dor_gdl = gdl_for(dor)

    win_types = group_by_type(win, win_det, win_gdl)
    dor_types = group_by_type(dor, dor_det, dor_gdl)
    print(f"Window types: {len(win_types)}  |  Door types: {len(dor_types)}", file=sys.stderr)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    build_title_block(wb)
    if win_types:
        build_sheet(wb, "External Windows", win_types)
    if dor_types:
        build_sheet(wb, "External Doors", dor_types)

    out = OUT_PATH
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print(f"\nWrote {out}", file=sys.stderr)
    print(f"Open with:\n  open '{out}'", file=sys.stderr)


if __name__ == "__main__":
    main()
